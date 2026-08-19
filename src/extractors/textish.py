"""Thin local parsers for common localization-ish formats. No network."""

from __future__ import annotations

import csv
import html
import io
import json
import re
from pathlib import Path
from typing import Any, Callable
from xml.etree import ElementTree as ET

from ..game_text import should_translate_string

TranslateFn = Callable[[list[str]], list[str]]


def _collect_json_strings(obj: Any, game_mode: bool, out: list[str]) -> None:
    if isinstance(obj, dict):
        for _k, v in obj.items():
            _collect_json_strings(v, game_mode, out)
    elif isinstance(obj, list):
        for v in obj:
            _collect_json_strings(v, game_mode, out)
    elif isinstance(obj, str) and should_translate_string(obj, game_mode):
        out.append(obj)


def _apply_json_strings(obj: Any, mapping: dict[str, str], game_mode: bool) -> Any:
    if isinstance(obj, dict):
        return {k: _apply_json_strings(v, mapping, game_mode) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_apply_json_strings(v, mapping, game_mode) for v in obj]
    if isinstance(obj, str) and should_translate_string(obj, game_mode):
        return mapping.get(obj, obj)
    return obj


def translate_json(path: Path, translate: TranslateFn, game_mode: bool) -> str:
    data = json.loads(path.read_text(encoding="utf-8"))
    strings: list[str] = []
    _collect_json_strings(data, game_mode, strings)
    if strings:
        mapped = dict(zip(strings, translate(strings)))
        data = _apply_json_strings(data, mapped, game_mode)
    return json.dumps(data, ensure_ascii=False, indent=2) + "\n"


def translate_yaml(path: Path, translate: TranslateFn, game_mode: bool) -> str:
    import yaml

    data = yaml.safe_load(path.read_text(encoding="utf-8"))
    strings: list[str] = []
    _collect_json_strings(data, game_mode, strings)
    if strings:
        mapped = dict(zip(strings, translate(strings)))
        data = _apply_json_strings(data, mapped, game_mode)
    return yaml.safe_dump(data, allow_unicode=True, sort_keys=False)


def translate_csv(path: Path, translate: TranslateFn, game_mode: bool, dialect: str = "excel") -> str:
    text = path.read_text(encoding="utf-8-sig")
    reader = csv.reader(io.StringIO(text), dialect=dialect)
    rows = [list(row) for row in reader]
    flat: list[str] = []
    coords: list[tuple[int, int]] = []
    for r, row in enumerate(rows):
        for c, cell in enumerate(row):
            if should_translate_string(cell, game_mode):
                coords.append((r, c))
                flat.append(cell)
    if flat:
        mapped = dict(zip(flat, translate(flat)))
        for r, c in coords:
            rows[r][c] = mapped.get(rows[r][c], rows[r][c])
    buf = io.StringIO()
    writer = csv.writer(buf, dialect=dialect, lineterminator="\n")
    writer.writerows(rows)
    return buf.getvalue()


def translate_tsv(path: Path, translate: TranslateFn, game_mode: bool) -> str:
    return translate_csv(path, translate, game_mode, dialect="excel-tab")


_PO_RE = re.compile(
    r'(msgid\s+"(?:[^"\\]|\\.)*"\s*msgstr\s+")((?:[^"\\]|\\.)*)(")',
    re.S,
)


def translate_po(path: Path, translate: TranslateFn, game_mode: bool) -> str:
    text = path.read_text(encoding="utf-8")
    # msgid is source; we fill msgstr. Never change msgid / keys.
    blocks = list(re.finditer(r'^msgid "(.*)"\nmsgstr "(.*)"', text, re.M))
    sources = []
    for m in blocks:
        src = bytes(m.group(1), "utf-8").decode("unicode_escape")
        dst = bytes(m.group(2), "utf-8").decode("unicode_escape")
        pick = dst if dst.strip() else src
        if should_translate_string(pick, game_mode) or (not dst.strip() and should_translate_string(src, False)):
            sources.append(pick if dst.strip() else src)
        else:
            sources.append("")
    to_send = [s for s in sources if s]
    mapped = dict(zip(to_send, translate(to_send))) if to_send else {}

    def repl(m: re.Match[str]) -> str:
        src = bytes(m.group(1), "utf-8").decode("unicode_escape")
        dst = bytes(m.group(2), "utf-8").decode("unicode_escape")
        pick = dst if dst.strip() else src
        new = mapped.get(pick, dst)
        escaped = (
            new.replace("\\", "\\\\")
            .replace('"', '\\"')
            .replace("\n", "\\n")
        )
        return f'msgid "{m.group(1)}"\nmsgstr "{escaped}"'

    return re.sub(r'^msgid "(.*)"\nmsgstr "(.*)"', repl, text, flags=re.M)


def _local(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def translate_xliff(path: Path, translate: TranslateFn, game_mode: bool) -> str:
    tree = ET.parse(path)
    root = tree.getroot()
    sources: list[str] = []
    nodes: list[ET.Element] = []
    for el in root.iter():
        if _local(el.tag) != "source":
            continue
        src = (el.text or "")
        parent = None
        # find sibling target under same parent by walking again is hard; store source nodes
        if src.strip() and should_translate_string(src, game_mode):
            sources.append(src)
            nodes.append(el)
    mapped = dict(zip(sources, translate(sources))) if sources else {}
    for parent in root.iter():
        kids = list(parent)
        source_el = next((c for c in kids if _local(c.tag) == "source"), None)
        if source_el is None:
            continue
        src = source_el.text or ""
        if src not in mapped:
            continue
        target_el = next((c for c in kids if _local(c.tag) == "target"), None)
        if target_el is None:
            target_el = ET.SubElement(parent, "target")
        target_el.text = mapped[src]
    ET.register_namespace("", root.tag.split("}")[0].strip("{") if root.tag.startswith("{") else "")
    return ET.tostring(root, encoding="unicode")


_TAG_SPLIT = re.compile(r"(<[^>]+>)")


def translate_html(path: Path, translate: TranslateFn, game_mode: bool) -> str:
    raw = path.read_text(encoding="utf-8")
    parts = _TAG_SPLIT.split(raw)
    skip = False
    idxs: list[int] = []
    texts: list[str] = []
    for i, part in enumerate(parts):
        if part.startswith("<"):
            low = part.lower()
            if re.match(r"<(script|style)\b", low):
                skip = True
            if re.match(r"</(script|style)>", low):
                skip = False
            continue
        if skip:
            continue
        if should_translate_string(part, game_mode):
            idxs.append(i)
            texts.append(part)
    if texts:
        mapped = dict(zip(texts, translate(texts)))
        for i in idxs:
            parts[i] = mapped.get(parts[i], parts[i])
    return "".join(parts)


_SRT_BLOCK = re.compile(
    r"(\d+\s*\n\d{2}:\d{2}:\d{2},\d{3}\s*-->\s*\d{2}:\d{2}:\d{2},\d{3}\s*\n)([\s\S]*?)(?=\n\n|\Z)"
)


def translate_srt(path: Path, translate: TranslateFn, game_mode: bool) -> str:
    text = path.read_text(encoding="utf-8")
    blocks = list(_SRT_BLOCK.finditer(text))
    bodies = []
    for m in blocks:
        body = m.group(2).strip("\n")
        bodies.append(body if should_translate_string(body, game_mode) or (body and not game_mode) else "")
    to_send = [b for b in bodies if b]
    mapped = dict(zip(to_send, translate(to_send))) if to_send else {}

    def repl(m: re.Match[str]) -> str:
        body = m.group(2).strip("\n")
        new = mapped.get(body, body)
        return m.group(1) + new + "\n"

    return _SRT_BLOCK.sub(repl, text)


def translate_vtt(path: Path, translate: TranslateFn, game_mode: bool) -> str:
    text = path.read_text(encoding="utf-8")
    cue_re = re.compile(
        r"((?:[\w-]+\n)?\d{2}:\d{2}:\d{2}\.\d{3}\s*-->\s*\d{2}:\d{2}:\d{2}\.\d{3}[^\n]*\n)([\s\S]*?)(?=\n\n|\Z)"
    )
    bodies = []
    matches = list(cue_re.finditer(text))
    for m in matches:
        body = m.group(2).strip("\n")
        if body.upper().startswith("NOTE") or body.strip() in {"STYLE", "REGION"}:
            bodies.append("")
        elif should_translate_string(body, game_mode) or (body and not game_mode):
            bodies.append(body)
        else:
            bodies.append("")
    to_send = [b for b in bodies if b]
    mapped = dict(zip(to_send, translate(to_send))) if to_send else {}

    def repl(m: re.Match[str]) -> str:
        body = m.group(2).strip("\n")
        new = mapped.get(body, body)
        return m.group(1) + new + "\n"

    return cue_re.sub(repl, text)


def translate_xlsx(path: Path, dest: Path, translate: TranslateFn, game_mode: bool) -> Path:
    from openpyxl import load_workbook

    wb = load_workbook(path)
    cells: list[str] = []
    coords: list[tuple[Any, Any]] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if not isinstance(val, str):
                    continue
                if val.startswith("="):
                    continue
                if should_translate_string(val, game_mode):
                    coords.append((cell, val))
                    cells.append(val)
    if cells:
        mapped = dict(zip(cells, translate(cells)))
        for cell, old in coords:
            cell.value = mapped.get(old, old)
    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)
    return dest


def unescape(text: str) -> str:
    return html.unescape(text)
